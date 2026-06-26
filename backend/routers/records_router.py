import base64
import csv
import io
import json
import re
from datetime import date, datetime

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from sqlalchemy.orm import Session
from sqlalchemy import DateTime, String, cast, or_, func
from typing import Optional
from backend.backup import backup_status, run_database_backup
from backend.database import get_db, engine
from backend.models import User, Client, BrokerContact, Property, AuditLog, PendingApproval, AppSetting
from backend.schemas import RecordCreate, RecordUpdate, ApprovalAction, MatchRequest, WorkflowAction
from backend.auth import get_current_user, require_permission, has_permission, normalize_role
from crm_core.attendance import calculate_attendance, normalize_status, parse_time
from crm_core.constants import (
    EXPENSE_CATEGORIES,
    normalize_availability_status,
    normalize_contact_role,
    DEAL_STAGES,
    DEAL_PRIORITIES,
    STAGE_PROBABILITY,
    DEAL_TABLES,
    COMMON_AREAS as PHASE1_AREAS,
    FACILITY_OPTIONS as PHASE1_FACILITIES,
    FLOOR_OPTIONS as PHASE1_FLOORS,
    PROPERTY_TYPE_OPTIONS as PHASE1_PROPERTY_TYPES,
    MEASUREMENT_UNIT_OPTIONS as PHASE1_MEASUREMENT_UNITS,
    CLOSED_AVAILABILITY_ARCHIVES,
    ADMIN_ROLES,
)
from crm_core.date_utils import DateUtils
from crm_core.ecosystem import collect_ecosystem_health
from crm_core.formatters import parse_currency
from crm_core.matching import best_matches
from crm_core.validators import PhoneValidator

router = APIRouter(prefix="/api/records", tags=["Records"])

PHASE1_TABLES = {"rent_requirements", "rent_availability", "sale_requirements", "sale_availability"}
PHASE1_LIST_LIMIT = 5000
READ_ONLY_TABLES = {"rented_properties", "sold_properties", "wf_instances", "wf_approvals", "wf_sla_log", "wf_audit_log"}

ALLOWED_TABLES = [
    "rent_requirements", "rent_availability",
    "rented_properties",
    "sale_requirements", "sale_availability",
    "sold_properties",
    "clients", "broker_contacts", "properties", "employees",
    "income_transactions", "expense_transactions",
    "attendance", "salary_payments",
    "sf_employees", "sf_positions", "sf_performance_goals",
    "sf_must_win_battles", "sf_kpis", "sf_learning", "sf_recruiting",
    "sf_compensation", "sf_onboarding",
    "wf_workflows", "wf_workflow_steps", "wf_instances", "wf_tasks",
    "wf_approvals", "wf_notifications", "wf_sla_log", "wf_audit_log",
]

GLOBAL_SEARCH_TABLES = [
    "rent_requirements",
    "rent_availability",
    "rented_properties",
    "sale_requirements",
    "sale_availability",
    "sold_properties",
    "clients",
    "broker_contacts",
    "properties",
    "employees",
    "income_transactions",
    "expense_transactions",
    "attendance",
    "salary_payments",
    "sf_employees",
    "sf_positions",
    "sf_performance_goals",
    "sf_must_win_battles",
    "sf_kpis",
    "sf_learning",
    "sf_recruiting",
    "sf_compensation",
    "sf_onboarding",
    "wf_workflows",
    "wf_workflow_steps",
    "wf_instances",
    "wf_tasks",
    "wf_approvals",
    "wf_notifications",
    "wf_sla_log",
    "wf_audit_log",
]

TABLE_LABELS = {
    "rent_requirements": "Rent Requirement",
    "rent_availability": "Rent Availability",
    "rented_properties": "Rented Property",
    "sale_requirements": "Sale Requirement",
    "sale_availability": "Sale Availability",
    "sold_properties": "Sold Property",
    "clients": "Client",
    "broker_contacts": "Brocker Contact",
    "properties": "Property",
    "employees": "Employee",
    "income_transactions": "Income Transaction",
    "expense_transactions": "Expense Transaction",
    "attendance": "Attendance",
    "salary_payments": "Salary Payment",
    "sf_employees": "SF Employee",
    "sf_positions": "SF Position",
    "sf_performance_goals": "SF Performance Goal",
    "sf_must_win_battles": "SF Must Win Battle",
    "sf_kpis": "SF KPI",
    "sf_learning": "SF Learning",
    "sf_recruiting": "SF Recruiting",
    "sf_compensation": "SF Compensation",
    "sf_onboarding": "SF Onboarding",
    "wf_workflows": "Workflow Definition",
    "wf_workflow_steps": "Workflow Step",
    "wf_instances": "Workflow Instance",
    "wf_tasks": "Workflow Task",
    "wf_approvals": "Workflow Approval",
    "wf_notifications": "Workflow Notification",
    "wf_sla_log": "Workflow SLA Log",
    "wf_audit_log": "Workflow Audit Log",
}

GLOBAL_SEARCH_HIDDEN_COLUMNS = {"password_hash"}
GLOBAL_SEARCH_EXCLUDED_COLUMNS = {"password_hash"}

TABLE_PERMISSIONS = {
    "rent_requirements": ("rent", "rent_view"),
    "rent_availability": ("rent", "rent_view"),
    "rented_properties": ("rent", "rent_view"),
    "sale_requirements": ("sale", "sale_view"),
    "sale_availability": ("sale", "sale_view"),
    "sold_properties": ("sale", "sale_view"),
    "clients": ("clients", None),
    "broker_contacts": ("clients", None),
    "properties": ("properties", None),
    "employees": ("employees", "employees_view"),
    "attendance": ("employees", "employees_view"),
    "salary_payments": ("employees", "employees_view"),
    "income_transactions": ("financial", "financial_view"),
    "expense_transactions": ("financial", "financial_view"),
    "sf_employees": ("successfactors", "sf_view"),
    "sf_positions": ("successfactors", "sf_view"),
    "sf_performance_goals": ("successfactors", "sf_view"),
    "sf_must_win_battles": ("successfactors", "sf_view"),
    "sf_kpis": ("successfactors", "sf_view"),
    "sf_learning": ("successfactors", "sf_view"),
    "sf_recruiting": ("successfactors", "sf_view"),
    "sf_compensation": ("successfactors", "sf_view"),
    "sf_onboarding": ("successfactors", "sf_view"),
    "wf_workflows": ("workflow", "wf_view"),
    "wf_workflow_steps": ("workflow", "wf_view"),
    "wf_instances": ("workflow", "wf_view"),
    "wf_tasks": ("workflow", "wf_view"),
    "wf_approvals": ("workflow", "wf_view"),
    "wf_notifications": ("workflow", "wf_view"),
    "wf_sla_log": ("workflow", "wf_view"),
    "wf_audit_log": ("workflow", "wf_view"),
}


PARENT_CHILD_TABLES = {
    "employees": (("attendance", "employee_id"), ("salary_payments", "employee_id")),
    "sf_employees": (
        ("sf_performance_goals", "employee_id"),
        ("sf_learning", "employee_id"),
        ("sf_compensation", "employee_id"),
        ("sf_onboarding", "employee_id"),
    ),
    "wf_workflows": (("wf_workflow_steps", "workflow_id"), ("wf_instances", "workflow_id")),
    "wf_instances": (("wf_tasks", "instance_id"), ("wf_sla_log", "instance_id")),
    "wf_tasks": (("wf_approvals", "task_id"), ("wf_sla_log", "task_id")),
}

CONTACT_FIELDS = {
    "rent_requirements": "contact_phone",
    "rent_availability": "owner_phone",
    "sale_requirements": "contact_phone",
    "sale_availability": "owner_phone",
    "rented_properties": "owner_phone",
    "sold_properties": "owner_phone",
    "clients": "phone",
    "broker_contacts": "contact",
    "properties": "owner_contact",
    "employees": "phone",
}

PHONE_ALIASES = {
    "rent_requirements": ("contact_phone", "contact", "owner_phone", "phone"),
    "rent_availability": ("owner_phone", "contact_phone", "contact", "owner_contact", "phone"),
    "sale_requirements": ("contact_phone", "contact", "owner_phone", "phone"),
    "sale_availability": ("owner_phone", "contact_phone", "contact", "owner_contact", "phone"),
    "rented_properties": ("owner_phone", "contact_phone", "contact", "owner_contact", "phone"),
    "sold_properties": ("owner_phone", "contact_phone", "contact", "owner_contact", "phone"),
    "clients": ("phone",),
    "broker_contacts": ("contact", "phone", "contact_phone"),
    "properties": ("owner_contact", "owner_phone", "contact_phone", "contact"),
    "employees": ("phone",),
}

PHASE1_PAYLOAD_ALIASES = {
    "contact",
    "contact_phone",
    "owner_phone",
    "owner_contact",
    "phone",
    "contact_person",
    "contact_status",
    "client_status",
    "client_broker",
    "budget_min",
    "budget_max",
}
MEASUREMENT_UNITS = {
    "sq ft": "Sq Ft",
    "sqft": "Sq Ft",
    "sqr ft": "Sq Ft",
    "square feet": "Sq Ft",
    "square ft": "Sq Ft",
    "yards": "Yards",
    "yard": "Yards",
}

REQUIRED_FIELDS = {
    "rent_requirements": ("date", "client_name", "contact_phone", "property_requires", "size", "floor", "location"),
    "rent_availability": ("date", "owner_name", "owner_phone", "property_availability", "size", "floor", "location", "monthly_rent"),
    "sale_requirements": ("date", "client_name", "contact_phone", "property_requires", "size", "floor", "budget", "location"),
    "sale_availability": ("date", "owner_name", "owner_phone", "property_availability", "size", "floor", "demand", "location"),
    "clients": ("client_name", "phone"),
    "broker_contacts": ("name", "contact"),
    "properties": ("title", "property_type", "location"),
    "employees": ("employee_id", "full_name", "phone"),
    "income_transactions": ("transaction_date", "income_type", "amount"),
    "expense_transactions": ("transaction_date", "expense_category", "amount"),
    "sf_employees": ("full_name", "department", "job_title"),
    "sf_positions": ("position_title",),
    "sf_performance_goals": ("employee_name", "goal_title"),
    "sf_must_win_battles": ("battle_title", "owner_name"),
    "sf_kpis": ("kpi_name",),
    "sf_learning": ("employee_name", "course_title"),
    "sf_recruiting": ("job_title",),
    "sf_compensation": ("employee_name", "base_salary"),
    "sf_onboarding": ("employee_name", "task_title"),
    "wf_workflows": ("workflow_name",),
    "wf_workflow_steps": ("step_name",),
    "wf_tasks": ("assigned_to",),
    "wf_notifications": ("recipient", "subject"),
}

DATE_FIELDS = {
    "date", "transaction_date", "hire_date", "payment_date", "next_follow_up", "last_contacted",
    "due_date", "start_date", "end_date", "open_date", "close_date", "assigned_date",
    "completion_date", "effective_date", "review_date", "due_at", "assigned_at",
    "completed_at",
}
MONEY_FIELDS = {
    "budget", "monthly_rent", "demand", "deposit", "maintenance_charge", "amount",
    "base_salary", "bonus", "deductions", "net_salary", "sale_price", "expected_close_value",
    "allowances", "total_compensation", "target_value", "current_value", "actual_value",
    "progress_pct", "weight_pct", "achievement_pct", "score", "actual_hours",
    "worked_minutes", "late_minutes", "early_leave_minutes", "overtime_minutes",
}
TIME_FIELDS = {"check_in", "check_out", "scheduled_start", "scheduled_end"}
TEXT_FILTER_FIELDS = {
    "rent_requirements": ("client_name", "contact", "contact_phone", "contact_person", "property_requires", "location", "facilities", "remarks"),
    "rent_availability": ("owner_name", "contact", "contact_phone", "owner_phone", "property_availability", "location", "facilities", "remarks"),
    "rented_properties": ("owner_name", "contact", "contact_phone", "owner_phone", "property_availability", "location", "facilities", "remarks", "closed_status"),
    "sale_requirements": ("client_name", "contact", "contact_phone", "contact_person", "property_requires", "location", "facilities", "remarks"),
    "sale_availability": ("owner_name", "contact", "contact_phone", "owner_phone", "property_availability", "location", "facilities", "remarks"),
    "sold_properties": ("owner_name", "contact", "contact_phone", "owner_phone", "property_availability", "location", "facilities", "remarks", "closed_status"),
    "clients": ("client_name", "phone", "email", "address", "client_type", "notes"),
    "broker_contacts": ("name", "contact", "area", "office_address", "home_address", "remarks"),
    "properties": ("property_code", "title", "property_type", "owner_name", "owner_contact", "location", "facilities"),
    "employees": ("employee_id", "full_name", "phone", "email", "position", "department"),
    "income_transactions": ("income_type", "tenant_name", "description", "receipt_no", "payment_method"),
    "expense_transactions": ("expense_category", "vendor_name", "description", "invoice_no", "payment_method"),
    "attendance": ("employee_id", "date", "check_in", "check_out", "status", "notes"),
    "salary_payments": ("employee_id", "payment_date", "month", "year", "payment_method", "notes"),
    "sf_employees": ("sf_employee_id", "full_name", "email", "department", "job_title", "manager_name", "employment_status", "location", "cost_center", "notes"),
    "sf_positions": ("position_code", "position_title", "department", "location", "status", "reports_to"),
    "sf_performance_goals": ("employee_name", "goal_title", "goal_description", "review_period", "status", "rating", "notes"),
    "sf_must_win_battles": ("battle_code", "battle_title", "owner_name", "department", "objective", "priority", "status", "business_impact", "risks", "notes"),
    "sf_kpis": ("kpi_code", "kpi_name", "employee_name", "owner_name", "department", "category", "period", "unit", "status", "notes"),
    "sf_learning": ("employee_name", "course_title", "course_code", "category", "instructor", "status", "notes"),
    "sf_recruiting": ("job_requisition_id", "job_title", "department", "location", "hiring_manager", "recruiter", "status", "notes"),
    "sf_compensation": ("employee_name", "currency", "review_cycle", "approved_by", "status", "notes"),
    "sf_onboarding": ("employee_name", "task_title", "task_category", "assigned_to", "status", "priority", "notes"),
    "wf_workflows": ("workflow_name", "workflow_type", "description", "trigger_event", "status"),
    "wf_workflow_steps": ("step_name", "step_type", "assignee_role", "assignee_name", "action_on_approve", "action_on_reject", "condition_field", "condition_value"),
    "wf_instances": ("workflow_name", "reference_table", "initiated_by", "current_assignee", "status", "priority", "notes"),
    "wf_tasks": ("workflow_name", "step_name", "assigned_to", "action_taken", "comments", "status", "priority", "reference_table"),
    "wf_approvals": ("workflow_name", "approval_type", "requested_by", "reviewed_by", "decision", "comments", "status"),
    "wf_notifications": ("recipient", "subject", "body", "channel", "status"),
    "wf_sla_log": ("breached", "logged_at"),
    "wf_audit_log": ("action", "performed_by", "reference_table", "old_value", "new_value", "ip_address", "session_id"),
}
KEYWORD_EXCLUDED_COLUMNS = {"password_hash", "is_deleted", "deleted_by", "deleted_at"}
SORT_DIRECTIONS = {"asc", "desc"}
DEFAULT_TABLE_SORTS = {
    "broker_contacts": ("area", "asc"),
}


def get_table_model(table: str):
    from backend.models import (
        RentRequirement, RentAvailability,
        SaleRequirement, SaleAvailability, RentedProperty, SoldProperty,
        IncomeTransaction, ExpenseTransaction,
        Client, BrokerContact, Property, Employee,
        Attendance, SalaryPayment,
        SFEmployee, SFPosition, SFPerformanceGoal, SFMustWinBattle, SFKPI,
        SFLearning, SFRecruiting, SFCompensation, SFOnboarding,
        WFWorkflow, WFWorkflowStep, WFInstance, WFTask, WFApproval,
        WFNotification, WFSlaLog, WFAuditLog,
    )
    mapping = {
        "rent_requirements": RentRequirement,
        "rent_availability": RentAvailability,
        "rented_properties": RentedProperty,
        "sale_requirements": SaleRequirement,
        "sale_availability": SaleAvailability,
        "sold_properties": SoldProperty,
        "income_transactions": IncomeTransaction,
        "expense_transactions": ExpenseTransaction,
        "clients": Client,
        "broker_contacts": BrokerContact,
        "properties": Property,
        "employees": Employee,
        "attendance": Attendance,
        "salary_payments": SalaryPayment,
        "sf_employees": SFEmployee,
        "sf_positions": SFPosition,
        "sf_performance_goals": SFPerformanceGoal,
        "sf_must_win_battles": SFMustWinBattle,
        "sf_kpis": SFKPI,
        "sf_learning": SFLearning,
        "sf_recruiting": SFRecruiting,
        "sf_compensation": SFCompensation,
        "sf_onboarding": SFOnboarding,
        "wf_workflows": WFWorkflow,
        "wf_workflow_steps": WFWorkflowStep,
        "wf_instances": WFInstance,
        "wf_tasks": WFTask,
        "wf_approvals": WFApproval,
        "wf_notifications": WFNotification,
        "wf_sla_log": WFSlaLog,
        "wf_audit_log": WFAuditLog,
    }
    return mapping.get(table)


def is_admin(user: User) -> bool:
    return normalize_role(user.role) in ADMIN_ROLES


def setting_value(db: Session, key: str, default: str = "") -> str:
    row = db.query(AppSetting).filter(AppSetting.key == key).first()
    return str(row.value) if row and row.value is not None else default


def normalize_setting_list(value: object, defaults: list[str]) -> list[str]:
    if isinstance(value, (list, tuple)):
        values = [str(item).strip() for item in value if str(item).strip()]
    else:
        raw = str(value or "")
        values = [item.strip() for item in raw.replace(",", "\n").splitlines() if item.strip()]
    if len(values) == 1 and defaults:
        packed = normalize_text(values[0])
        unpacked = [item for item in defaults if normalize_text(item) in packed]
        if len(unpacked) > 1:
            return unpacked
    return values or defaults


def setting_list(db: Session, key: str, defaults: list[str]) -> list[str]:
    return normalize_setting_list(setting_value(db, key, "\n".join(defaults)), defaults)


def setting_list_text(value: object, defaults: list[str]) -> str:
    return "\n".join(normalize_setting_list(value, defaults))


def set_setting(db: Session, key: str, value: str) -> None:
    row = db.query(AppSetting).filter(AppSetting.key == key).first()
    if row:
        row.value = value
    else:
        db.add(AppSetting(key=key, value=value))


def normalize_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def multi_value_tokens(value: object) -> set[str]:
    return {
        normalize_text(part)
        for part in re.split(r"[,;|\n]+", str(value or ""))
        if normalize_text(part)
    }


def parse_facilities(value: object, options: list[str]) -> set[str]:
    aliases = {
        "parking": "Car Parking",
        "car park": "Car Parking",
        "bike": "Bike Parking",
        "cctv": "CCTV Camera",
        "camera": "CCTV Camera",
        "electricity 24/7": "Light 24/7",
        "light 24/7": "Light 24/7",
        "light with loadshading": "Light With Loadshedding",
        "load shedding": "Light With Loadshedding",
        "loadshading": "Light With Loadshedding",
    }
    lookup = {normalize_text(option): option for option in options}
    selected: set[str] = set()
    for token in re.split(r"[,;|\n]+", str(value or "")):
        normalized = aliases.get(normalize_text(token), token)
        key = normalize_text(normalized)
        if key in lookup:
            selected.add(lookup[key])
    return selected


def nearby_location(left: str, right: str) -> bool:
    groups = [
        {"gizri", "dha", "defence", "dha phase 4", "dha phase 5", "dha phase 6", "zamzama", "clifton", "boat basin", "sea view", "marina"},
        {"clifton", "clifton block 1", "clifton block 2", "clifton block 3", "clifton block 4", "clifton block 5", "clifton block 6", "clifton block 7", "clifton block 8", "clifton block 9", "boat basin"},
        {"pechs", "tariq road", "bahadurabad", "kda scheme"},
        {"north nazimabad", "nazimabad", "fb area", "hyderi", "water pump"},
    ]
    return any(left in group and right in group for group in groups)


def can_read_table(user: User, table: str) -> bool:
    # Phase 1 deal desk records are office-wide. Staff should see every entry
    # in these four sections, regardless of who created it.
    if table in PHASE1_TABLES and normalize_role(user.role) in {"Super Admin", "Admin", "Manager", "Staff", "Viewer"}:
        return True
    edit_perm, view_perm = TABLE_PERMISSIONS.get(table, (None, None))
    if edit_perm and has_permission(user.role, edit_perm):
        return True
    if view_perm and has_permission(user.role, view_perm):
        return True
    return False


def can_write_table(user: User, table: str) -> bool:
    if table in READ_ONLY_TABLES:
        return False
    if table in PHASE1_TABLES and normalize_role(user.role) in {"Super Admin", "Admin", "Manager", "Staff"}:
        return True
    edit_perm, _view_perm = TABLE_PERMISSIONS.get(table, (None, None))
    return bool(edit_perm and has_permission(user.role, edit_perm))


def require_table_read(user: User, table: str) -> None:
    if not can_read_table(user, table):
        raise HTTPException(status_code=403, detail=f"You do not have access to {TABLE_LABELS.get(table, table)}")


def require_table_write(user: User, table: str) -> None:
    if not can_write_table(user, table):
        raise HTTPException(status_code=403, detail=f"You cannot change {TABLE_LABELS.get(table, table)}")


def first_present(mapping: dict, *keys: str):
    for key in keys:
        value = mapping.get(key)
        if value not in (None, ""):
            return value
    return ""


def serialize_record(record):
    data = {c.name: getattr(record, c.name) for c in record.__table__.columns if c.name not in GLOBAL_SEARCH_HIDDEN_COLUMNS}
    table = getattr(record, "__tablename__", "") or getattr(getattr(record, "__table__", None), "name", "")
    if table in {"rent_requirements", "sale_requirements"}:
        phone = first_present(data, "contact_phone", "contact")
        data["contact_phone"] = phone
        data["contact"] = phone
        data["contact_person"] = first_present(data, "contact_person", "client_name")
        data["contact_status"] = first_present(data, "client_status") or "Client"
    elif table in {"rent_availability", "sale_availability", "rented_properties", "sold_properties"}:
        phone = first_present(data, "owner_phone", "contact_phone", "contact")
        data["owner_phone"] = phone
        data["contact_phone"] = phone
        data["contact"] = phone
        data["contact_status"] = first_present(data, "client_broker") or "Owner"
        if "status" in data:
            try:
                data["status"] = normalize_availability_status(data.get("status"), "Available")
            except ValueError:
                data["status"] = data.get("status") or "Available"
    return data


def model_columns(model):
    return {c.name for c in model.__table__.columns}


def model_column_map(model):
    return {c.name: c for c in model.__table__.columns}


def is_datetime_column(column) -> bool:
    return isinstance(column.type, DateTime)


def timestamp_for_column(column):
    now = datetime.now()
    if is_datetime_column(column):
        return now
    return now.isoformat(timespec="seconds")


def coerce_datetime_columns(model, data: dict) -> None:
    for name, column in model_column_map(model).items():
        if name not in data or data[name] in (None, "") or not is_datetime_column(column):
            continue
        value = data[name]
        if isinstance(value, datetime):
            continue
        if isinstance(value, date):
            data[name] = datetime.combine(value, datetime.min.time())
            continue
        if isinstance(value, str):
            try:
                data[name] = datetime.fromisoformat(value.strip().replace("Z", "+00:00"))
                continue
            except ValueError:
                pass
        raise HTTPException(
            status_code=422,
            detail={
                "message": "Please fix the highlighted record fields.",
                "errors": [f"{name.replace('_', ' ').title()} must be a valid date/time"],
            },
        )


def keyword_search_columns(model, table: str) -> list[str]:
    configured = [col for col in TEXT_FILTER_FIELDS.get(table, ()) if hasattr(model, col)]
    all_columns = [
        c.name for c in model.__table__.columns
        if c.name not in KEYWORD_EXCLUDED_COLUMNS and hasattr(model, c.name)
    ]
    columns: list[str] = []
    for col in configured + all_columns:
        if col not in columns:
            columns.append(col)
    return columns


def split_text_filter_terms(value: object) -> list[str]:
    return [term.strip() for term in re.split(r"[,;]+", str(value or "")) if term.strip()]


def apply_column_text_filter(query, model, column: str, value: object):
    if not value or not hasattr(model, column):
        return query
    terms = split_text_filter_terms(value)
    if not terms:
        return query
    model_column = getattr(model, column)
    return query.filter(or_(*[cast(model_column, String).ilike(f"%{term}%") for term in terms]))


def normalize_phone(value: object) -> str:
    return PhoneValidator.normalize(value)


def parse_date_value(value: object) -> str:
    if value in (None, ""):
        return ""
    return DateUtils.store_date(value)


def phone_aliases_for_table(table: str) -> tuple[str, ...]:
    return PHONE_ALIASES.get(table, (CONTACT_FIELDS.get(table) or "",))


def payload_phone(table: str, data: dict) -> str:
    for field in phone_aliases_for_table(table):
        if field and data.get(field) not in (None, ""):
            return normalize_phone(data.get(field))
    return ""


def normalize_phase1_payload(table: str, payload: dict, columns: set[str]) -> dict:
    if table not in PHASE1_TABLES:
        return {k: v for k, v in payload.items() if k in columns and k != "id"}

    data = dict(payload)
    is_requirement = table in {"rent_requirements", "sale_requirements"}
    role_field = "client_status" if is_requirement else "client_broker"
    default_role = "Client" if is_requirement else "Owner"

    role_value = first_present(data, "contact_status", role_field, "client_status", "client_broker")
    if role_field in columns:
        data[role_field] = normalize_contact_role(role_value, default_role)

    phone_was_sent = any(field in data for field in phone_aliases_for_table(table))
    if phone_was_sent:
        phone = payload_phone(table, data)
        targets = ("contact", "contact_phone") if is_requirement else ("contact", "contact_phone", "owner_phone")
        for target in targets:
            if target in columns:
                data[target] = phone

    if is_requirement and "contact_person" in columns:
        data["contact_person"] = first_present(data, "contact_person", "client_name")

    if not is_requirement and "status" in data and "status" in columns:
        data["status"] = normalize_availability_status(data.get("status"), "Available")

    budget_was_sent = any(field in data for field in ("budget", "budget_max", "budget_min"))
    if "budget" in columns and budget_was_sent and data.get("budget") in (None, "", 0, "0"):
        data["budget"] = first_present(data, "budget_max", "budget_min", "budget")

    return {k: v for k, v in data.items() if k in columns and k != "id"}


def validate_record_payload(table: str, data: dict, *, creating: bool = False) -> None:
    errors: list[str] = []
    required = REQUIRED_FIELDS.get(table, ())
    if creating:
        for field in required:
            if str(data.get(field) or "").strip() == "":
                errors.append(f"{field.replace('_', ' ').title()} is required")
    else:
        for field in required:
            if field in data and str(data.get(field) or "").strip() == "":
                errors.append(f"{field.replace('_', ' ').title()} cannot be empty")

    for field in DATE_FIELDS & set(data):
        if data.get(field) in (None, ""):
            continue
        try:
            data[field] = parse_date_value(data[field])
        except ValueError as exc:
            errors.append(f"{field.replace('_', ' ').title()}: {exc}")

    for field in MONEY_FIELDS & set(data):
        value = data.get(field)
        if value in (None, ""):
            data[field] = 0
            continue
        number = parse_currency(value)
        if number is None:
            errors.append(f"{field.replace('_', ' ').title()} must be a number")
            continue
        if number < 0:
            errors.append(f"{field.replace('_', ' ').title()} cannot be negative")
        data[field] = number

    for field in TIME_FIELDS & set(data):
        value = data.get(field)
        if value in (None, ""):
            continue
        parsed = parse_time(value)
        if not parsed:
            errors.append(f"{field.replace('_', ' ').title()} must be a valid time")
            continue
        data[field] = parsed.strftime("%H:%M")

    if table == "attendance" and "status" in data:
        data["status"] = normalize_status(data.get("status"))

    if "measurement" in data and data.get("measurement") not in (None, ""):
        try:
            measurement = float(str(data.get("measurement")).replace(",", "").strip())
        except (TypeError, ValueError):
            errors.append("Measurement must be a number")
        else:
            if measurement < 0:
                errors.append("Measurement cannot be negative")
            data["measurement"] = str(int(measurement)) if measurement.is_integer() else str(measurement)

    if "measurement_unit" in data and data.get("measurement_unit") not in (None, ""):
        unit_key = normalize_text(data.get("measurement_unit"))
        if unit_key not in MEASUREMENT_UNITS:
            errors.append("Size must be Sq Ft or Yards")
        else:
            data["measurement_unit"] = MEASUREMENT_UNITS[unit_key]
    elif data.get("measurement") not in (None, "") and "measurement_unit" in data:
        data["measurement_unit"] = "Sq Ft"

    for phone_field in phone_aliases_for_table(table):
        if phone_field and phone_field in data:
            try:
                phone = PhoneValidator.validate_phone(data.get(phone_field))
            except ValueError as exc:
                errors.append(f"{phone_field.replace('_', ' ').title()}: {exc}")
                continue
            data[phone_field] = phone

    if table in {"rent_requirements", "sale_requirements"} and "client_status" in data:
        try:
            data["client_status"] = normalize_contact_role(data.get("client_status"), "Client")
        except ValueError as exc:
            errors.append(str(exc))
    if table in {"rent_availability", "sale_availability"}:
        if "client_broker" in data:
            try:
                data["client_broker"] = normalize_contact_role(data.get("client_broker"), "Owner")
            except ValueError as exc:
                errors.append(str(exc))
        if "status" in data:
            try:
                data["status"] = normalize_availability_status(data.get("status"), "Available")
            except ValueError as exc:
                errors.append(str(exc))

    if "workflow_stage" in data and data["workflow_stage"] not in (None, ""):
        data["workflow_stage"] = normalize_stage(str(data["workflow_stage"]))
    if "priority" in data and data["priority"] not in DEAL_PRIORITIES:
        errors.append(f"Priority must be one of {', '.join(DEAL_PRIORITIES)}")

    if errors:
        raise HTTPException(
            status_code=422,
            detail={"message": "Please fix the highlighted record fields.", "errors": errors},
        )


def clean_payload(model, payload: dict, user: Optional[User] = None, creating: bool = False, table: str = ""):
    column_map = model_column_map(model)
    columns = set(column_map)
    allowed_aliases = PHASE1_PAYLOAD_ALIASES if table in PHASE1_TABLES else set()
    unknown = sorted(k for k in payload.keys() if (k not in columns and k not in allowed_aliases) or k == "id")
    if unknown:
        raise HTTPException(status_code=400, detail=f"Unknown fields: {unknown}")
    try:
        data = normalize_phase1_payload(table, payload, columns)
    except ValueError as exc:
        raise HTTPException(
            status_code=422,
            detail={"message": "Please fix the highlighted record fields.", "errors": [str(exc)]},
        ) from exc
    if creating and user is not None and "created_by" in columns and not data.get("created_by"):
        data["created_by"] = user.username
    if creating and "created_at" in columns and not data.get("created_at"):
        data["created_at"] = timestamp_for_column(column_map["created_at"])
    if table == "attendance":
        data = calculate_attendance(data)
        data = {key: value for key, value in data.items() if key in columns}
    coerce_datetime_columns(model, data)
    return data


def record_value(record, *names: str):
    for name in names:
        if hasattr(record, name):
            value = getattr(record, name)
            if value not in (None, ""):
                return value
    return ""


def record_text(record, *names: str) -> str:
    return str(record_value(record, *names) or "").strip()


def record_number(record, *names: str) -> float:
    value = record_value(record, *names)
    number = parse_currency(value)
    return float(number or 0.0)


def contact_type(value, default: str) -> str:
    text = str(value or "").strip().lower()
    if text in {"b", "broker", "agent"}:
        return "Broker"
    if text in {"o", "owner", "seller"}:
        return "Owner"
    return default


def deal_client_contacts(table: str, record) -> list[dict[str, str]]:
    contacts: list[dict[str, str]] = []
    if table in {"rent_requirements", "sale_requirements"}:
        default_type = "Tenant" if table.startswith("rent") else "Buyer"
        contacts.append({
            "name": record_text(record, "client_name"),
            "phone": record_text(record, "contact_phone", "contact"),
            "email": record_text(record, "contact_email"),
            "type": contact_type(record_text(record, "client_status"), default_type),
        })
    elif table in {"rent_availability", "sale_availability"}:
        contacts.append({
            "name": record_text(record, "owner_name"),
            "phone": record_text(record, "owner_phone", "contact_phone", "contact"),
            "email": record_text(record, "contact_email"),
            "type": contact_type(record_text(record, "client_broker"), "Owner"),
        })
    for key in ("broker", "preferred_broker", "posted_by_broker", "posted_by", "client_broker"):
        broker = record_text(record, key)
        if broker and broker.lower() not in {"o", "b", "owner", "broker", "direct", "agent", "client"}:
            contacts.append({"name": broker, "phone": "", "email": "", "type": "Broker"})
    return contacts


def upsert_clients_from_deal(db: Session, table: str, record) -> None:
    if table not in DEAL_TABLES:
        return
    for contact in deal_client_contacts(table, record):
        name = contact["name"]
        if not name:
            continue
        phone = contact["phone"]
        email = contact["email"]
        client_type = contact["type"] or "Other"
        notes = f"Auto-synced from {TABLE_LABELS.get(table, table)} #{getattr(record, 'id', '')}".strip()
        existing = None
        if phone:
            existing = db.query(Client).filter(Client.phone == phone).first()
        if existing is None:
            existing = db.query(Client).filter(func.lower(Client.client_name) == name.lower()).first()
        if existing:
            if not existing.client_name:
                existing.client_name = name
            if phone and not existing.phone:
                existing.phone = phone
            if email and not existing.email:
                existing.email = email
            if client_type and not existing.client_type:
                existing.client_type = client_type
            if not existing.status:
                existing.status = "Active"
            if not existing.notes:
                existing.notes = notes
        else:
            db.add(Client(
                client_name=name,
                phone=phone,
                email=email,
                client_type=client_type,
                status="Active",
                notes=notes,
            ))


def property_match(db: Session, record, title: str, property_type: str):
    location = record_text(record, "location")
    owner_name = record_text(record, "owner_name")
    owner_contact = record_text(record, "owner_phone", "contact_phone", "contact")
    floor = record_text(record, "floor", "floor_no")

    def text_match(query, column, value: str):
        if not value:
            return query
        return query.filter(func.lower(func.coalesce(column, "")) == value.lower())

    if owner_contact and location:
        query = db.query(Property).filter(Property.owner_contact == owner_contact)
        query = text_match(query, Property.location, location)
        query = text_match(query, Property.property_type, property_type)
        floor_match = text_match(query, Property.floor, floor).first() if floor else None
        if floor_match:
            return floor_match
        found = query.first()
        if found:
            return found
    if owner_name and location:
        query = db.query(Property)
        query = text_match(query, Property.owner_name, owner_name)
        query = text_match(query, Property.location, location)
        query = text_match(query, Property.property_type, property_type)
        found = query.first()
        if found:
            return found
    if title and location:
        query = db.query(Property)
        query = text_match(query, Property.title, title)
        query = text_match(query, Property.location, location)
        query = text_match(query, Property.property_type, property_type)
        found = query.first()
        if found:
            return found
    if owner_contact and not location:
        return db.query(Property).filter(Property.owner_contact == owner_contact).first()
    return None


def availability_property_status(table: str, record) -> str:
    try:
        status = normalize_availability_status(record_text(record, "status"), "Available")
    except ValueError:
        status = record_text(record, "status") or "Available"
    stage = normalize_stage(record_text(record, "workflow_stage"))
    if stage == "Pending" and status == "Available":
        return "Pending"
    if table == "rent_availability" and status == "Sold":
        return "Available"
    if table == "sale_availability" and status == "Rented":
        return "Available"
    return status


def sync_property_from_availability(db: Session, table: str, record, status: str) -> None:
    if table not in {"rent_availability", "sale_availability"}:
        return
    property_type = record_text(record, "property_availability", "property_type")
    location = record_text(record, "location")
    if not property_type and not location:
        return
    title = f"{property_type or 'Property'} - {location or 'Location'}"
    area = " ".join(part for part in (record_text(record, "measurement"), record_text(record, "measurement_unit")) if part)
    data = {
        "title": title,
        "property_type": property_type,
        "status": status,
        "owner_name": record_text(record, "owner_name"),
        "owner_contact": record_text(record, "owner_phone", "contact_phone", "contact"),
        "location": location,
        "area": area,
        "floor": record_text(record, "floor", "floor_no"),
        "bedrooms": record_text(record, "size"),
        "maintenance_charge": record_number(record, "maintenance_charge"),
        "facilities": record_text(record, "facilities"),
        "description": record_text(record, "remarks", "description", "notes"),
    }
    if table.startswith("rent"):
        data["monthly_rent"] = record_number(record, "monthly_rent")
    elif table.startswith("sale"):
        data["sale_price"] = record_number(record, "demand", "asking_price")
    existing = property_match(db, record, title, property_type)
    if existing:
        for key, value in data.items():
            setattr(existing, key, value)
    else:
        db.add(Property(property_code=f"PROP{datetime.now().strftime('%Y%m%d%H%M%S')}", **data))


def archive_closed_availability_record(db: Session, table: str, record, username: str = "system"):
    rule = CLOSED_AVAILABILITY_ARCHIVES.get(table)
    if not rule or record is None:
        return None
    closed_status, archive_table, deal_type = rule
    try:
        status = normalize_availability_status(record_text(record, "status"), "Available")
    except ValueError:
        status = record_text(record, "status")
    if status != closed_status:
        return None

    archive_model = get_table_model(archive_table)
    if archive_model is None:
        return None

    now = datetime.now().isoformat(timespec="seconds")
    if hasattr(record, "workflow_stage"):
        record.workflow_stage = "Deal Done"
    if hasattr(record, "deal_probability"):
        record.deal_probability = 100
    if hasattr(record, "closed_at") and not getattr(record, "closed_at", None):
        record.closed_at = datetime.now()
    if hasattr(record, "is_deleted"):
        record.is_deleted = True
    if hasattr(record, "deleted_by"):
        record.deleted_by = username
    if hasattr(record, "deleted_at"):
        record.deleted_at = now

    archive = (
        db.query(archive_model)
        .filter(archive_model.source_table == table, archive_model.source_id == record.id)
        .first()
    )
    if not archive:
        archive = archive_model(source_table=table, source_id=record.id)
        db.add(archive)

    archive_columns = model_columns(archive_model)
    values = {
        "deal_type": deal_type,
        "closed_status": closed_status,
        "closed_at": record_text(record, "closed_at") or now,
        "archived_at": now,
        "archived_by": username,
        "workflow_stage": record_text(record, "workflow_stage") or "Deal Done",
        "deal_probability": getattr(record, "deal_probability", None) or 100,
        "original_payload": json.dumps(serialize_record(record), default=str, ensure_ascii=True),
    }
    for column in (
        "date", "owner_name", "owner_phone", "contact_phone", "contact",
        "property_availability", "size", "measurement", "measurement_unit",
        "monthly_rent", "demand", "deposit", "maintenance_charge", "floor",
        "location", "bedrooms", "bathrooms", "furnishing", "parking",
        "nearby_landmarks", "area_notes", "verification_status", "photo_paths",
        "facilities", "client_broker", "bachelor_family", "remarks", "persons",
        "building_name", "priority", "assigned_to", "expected_close_value",
        "approval_status", "created_by", "created_at",
    ):
        if hasattr(record, column):
            values[column] = getattr(record, column)
    for key, value in values.items():
        if key in archive_columns:
            setattr(archive, key, value)
    return archive


def sync_after_deal_save(db: Session, table: str, record, username: str = "system") -> None:
    if table not in DEAL_TABLES:
        return
    upsert_clients_from_deal(db, table, record)
    if table in {"rent_availability", "sale_availability"}:
        sync_property_from_availability(db, table, record, availability_property_status(table, record))
        archive_closed_availability_record(db, table, record, username)


def sync_all_deal_inventory(db: Session) -> dict[str, int]:
    """Backfill clients and property inventory from active Phase 1 deal rows."""
    synced_records = 0
    synced_properties = 0
    for table in DEAL_TABLES:
        model = get_table_model(table)
        if model is None:
            continue
        query = db.query(model)
        if hasattr(model, "is_deleted"):
            deleted_col = getattr(model, "is_deleted")
            query = query.filter(deleted_col.is_(False) | (deleted_col == 0) | deleted_col.is_(None))
        for record in query.all():
            upsert_clients_from_deal(db, table, record)
            if table in {"rent_availability", "sale_availability"}:
                sync_property_from_availability(db, table, record, availability_property_status(table, record))
                synced_properties += 1
            synced_records += 1
    return {"records": synced_records, "properties": synced_properties}


def normalize_stage(stage: Optional[str]) -> str:
    return stage if stage in DEAL_STAGES else "Lead"


def deal_row(table: str, record):
    name_col, contact_col, type_col, value_col = DEAL_TABLES[table]
    stage = normalize_stage(getattr(record, "workflow_stage", None))
    amount = getattr(record, "expected_close_value", None) or getattr(record, value_col, 0) or 0
    return {
        "table": table,
        "id": record.id,
        "name": getattr(record, name_col, "") or "",
        "contact": record_text(record, contact_col, "owner_phone", "contact_phone", "contact"),
        "location": getattr(record, "location", "") or "",
        "property_type": getattr(record, type_col, "") or "",
        "amount": float(amount or 0),
        "stage": stage,
        "priority": getattr(record, "priority", None) or "Medium",
        "next_follow_up": getattr(record, "next_follow_up", None),
        "assigned_to": getattr(record, "assigned_to", None),
        "probability": float(getattr(record, "deal_probability", None) or STAGE_PROBABILITY[stage]),
    }


def record_label(record) -> str:
    return (
        record_text(record, "client_name")
        or record_text(record, "owner_name")
        or record_text(record, "full_name")
        or record_text(record, "title")
        or record_text(record, "property_code")
        or f"#{getattr(record, 'id', '')}"
    )


def find_duplicate_records(
    db: Session,
    user: User,
    table: str,
    data: dict,
    *,
    exclude_id: int | None = None,
    max_per_table: int = 8,
) -> list[dict]:
    if table not in CONTACT_FIELDS:
        return []
    phone = payload_phone(table, data)
    if not phone:
        return []
    duplicates: list[dict] = []
    for other_table in CONTACT_FIELDS:
        if other_table in ALLOWED_TABLES and not can_read_table(user, other_table):
            continue
        model = get_table_model(other_table)
        fields = [field for field in phone_aliases_for_table(other_table) if field and hasattr(model, field)] if model is not None else []
        if model is None or not fields:
            continue
        query = db.query(model).filter(or_(*[getattr(model, field).isnot(None) for field in fields]))
        if hasattr(model, "is_deleted"):
            deleted_col = getattr(model, "is_deleted")
            query = query.filter(deleted_col.is_(False) | (deleted_col == 0) | deleted_col.is_(None))
        rows = query.order_by(model.id.desc()).limit(2000).all()
        found = 0
        for row in rows:
            if other_table == table and exclude_id and getattr(row, "id", None) == exclude_id:
                continue
            if not any(normalize_phone(getattr(row, field, "")) == phone for field in fields):
                continue
            duplicates.append({
                "table": other_table,
                "table_label": TABLE_LABELS.get(other_table, other_table),
                "id": getattr(row, "id", None),
                "name": record_label(row),
                "phone": phone,
                "location": record_text(row, "location", "address"),
            })
            found += 1
            if found >= max_per_table:
                break
    return duplicates


def audit_snapshot(record) -> dict:
    return serialize_record(record) if record is not None else {}


def audit_summary(table: str, record_id: int | None, action: str, before: dict | None, after: dict | None) -> str:
    label = TABLE_LABELS.get(table, table)
    if action == "create":
        return f"Created {label} #{record_id}"
    if action == "delete":
        return f"Deleted {label} #{record_id}"
    if action == "workflow":
        return f"Workflow changed for {label} #{record_id}"
    if action == "approve":
        return f"Approval changed for {label} #{record_id}"
    if before and after:
        changed = [key for key in sorted(after) if before.get(key) != after.get(key)]
        if changed:
            fields = ", ".join(changed[:6])
            suffix = "..." if len(changed) > 6 else ""
            return f"Updated {label} #{record_id}: {fields}{suffix}"
    return f"Updated {label} #{record_id}"


def write_audit_log(
    db: Session,
    user: User,
    table: str,
    record_id: int | None,
    action: str,
    *,
    before: dict | None = None,
    after: dict | None = None,
) -> None:
    details = {"before": before or {}, "after": after or {}}
    db.add(AuditLog(
        table_name=table,
        record_id=record_id,
        action=action,
        username=user.username,
        summary=audit_summary(table, record_id, action, before, after),
        details=json.dumps(details, default=str, ensure_ascii=True),
    ))


def date_column_for_table(model):
    for column_name in (
        "date", "transaction_date", "payment_date", "hire_date", "due_date",
        "start_date", "end_date", "open_date", "assigned_date", "effective_date",
        "review_date", "initiated_at", "due_at", "assigned_at", "requested_at",
        "sent_at", "logged_at", "performed_at", "created_at",
    ):
        if hasattr(model, column_name):
            return getattr(model, column_name)
    return None


def apply_list_filters(
    query,
    model,
    table: str,
    *,
    q: str = "",
    stage: str | None = None,
    status: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    field_filters: dict[str, object] | None = None,
    include_deleted: bool = False,
    deleted_only: bool = False,
):
    if hasattr(model, "is_deleted"):
        deleted_col = getattr(model, "is_deleted")
        if deleted_only:
            query = query.filter(deleted_col.is_(True) | (deleted_col == 1))
        elif not include_deleted:
            query = query.filter(deleted_col.is_(False) | (deleted_col == 0) | deleted_col.is_(None))
    closed_rule = CLOSED_AVAILABILITY_ARCHIVES.get(table)
    if closed_rule and not include_deleted and not deleted_only and hasattr(model, "status"):
        query = query.filter(func.lower(func.coalesce(getattr(model, "status"), "")) != closed_rule[0].lower())
    term = (q or "").strip()
    if term:
        pattern = f"%{term}%"
        columns = keyword_search_columns(model, table)
        if columns:
            query = query.filter(or_(*[cast(getattr(model, col), String).ilike(pattern) for col in columns]))
    if stage and hasattr(model, "workflow_stage"):
        query = query.filter(getattr(model, "workflow_stage") == stage)
    if status and hasattr(model, "status"):
        if table in {"rent_availability", "sale_availability"}:
            try:
                status = normalize_availability_status(status)
            except ValueError as exc:
                raise HTTPException(status_code=400, detail=str(exc))
        query = query.filter(getattr(model, "status") == status)
    for column, value in (field_filters or {}).items():
        query = apply_column_text_filter(query, model, column, value)
    date_col = date_column_for_table(model)
    if date_col is not None:
        if date_from:
            try:
                parsed_from = parse_date_value(date_from)
            except ValueError as exc:
                raise HTTPException(status_code=400, detail=f"date_from: {exc}")
            query = query.filter(func.date(date_col) >= parsed_from)
        if date_to:
            try:
                parsed_to = parse_date_value(date_to)
            except ValueError as exc:
                raise HTTPException(status_code=400, detail=f"date_to: {exc}")
            query = query.filter(func.date(date_col) <= parsed_to)
    return query


def require_audit_permission(user: User) -> None:
    if not has_permission(user.role, "users"):
        raise HTTPException(status_code=403, detail="Only admins can view audit logs")


def child_reference_summary(db: Session, table: str, record_id: int) -> list[str]:
    references: list[str] = []
    for child_table, child_column in PARENT_CHILD_TABLES.get(table, ()):
        child_model = get_table_model(child_table)
        if child_model is None or not hasattr(child_model, child_column):
            continue
        count = (
            db.query(child_model)
            .filter(getattr(child_model, child_column) == record_id)
            .count()
        )
        if count:
            references.append(f"{TABLE_LABELS.get(child_table, child_table)}: {count}")
    return references


def require_no_child_references(db: Session, table: str, record_id: int) -> None:
    references = child_reference_summary(db, table, record_id)
    if references:
        raise HTTPException(
            status_code=409,
            detail=(
                "Cannot delete this record because related records exist: "
                + "; ".join(references)
            ),
        )


# Specific routes must come before parameterized /{table} routes
@router.get("/search/global")
def global_search(
    q: str = Query("", min_length=1),
    limit: int = Query(100, le=500),
    table: str | None = Query(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    results = []
    term = q.strip()
    if not term:
        return {"ok": True, "query": q, "count": 0, "results": []}
    pattern = f"%{term}%"
    normalized_term = term.lower()
    requested_table = (table or "").strip().lower()
    if requested_table:
        if requested_table not in GLOBAL_SEARCH_TABLES:
            raise HTTPException(status_code=400, detail=f"Invalid search table: {requested_table}")
        target_tables = [requested_table]
    else:
        target_tables = list(GLOBAL_SEARCH_TABLES)
    for search_table in target_tables:
        if not can_read_table(user, search_table):
            continue
        try:
            model = get_table_model(search_table)
            if model is None:
                continue
            searchable_columns = [
                c.name for c in model.__table__.columns
                if c.name not in GLOBAL_SEARCH_EXCLUDED_COLUMNS and hasattr(model, c.name)
            ]
            return_columns = [c.name for c in model.__table__.columns if c.name not in GLOBAL_SEARCH_HIDDEN_COLUMNS]
            if not searchable_columns:
                continue
            source = TABLE_LABELS.get(search_table, search_table.replace("_", " ").title())
            source_text = f"{source} {search_table.replace('_', ' ')}".lower()
            query = db.query(model)
            if hasattr(model, "is_deleted"):
                query = query.filter(getattr(model, "is_deleted").is_(False) | (getattr(model, "is_deleted") == 0) | getattr(model, "is_deleted").is_(None))
            if normalized_term in source_text:
                records = query.order_by(getattr(model, "id").desc()).limit(limit).all()
            else:
                records = (
                    query.filter(
                        or_(
                            *[cast(getattr(model, col), String).ilike(pattern) for col in searchable_columns]
                        )
                    )
                    .order_by(getattr(model, "id").desc())
                    .limit(limit)
                    .all()
                )
            for r in records:
                fields = serialize_record(r)
                fields = {key: value for key, value in fields.items() if key in return_columns}
                matched_columns = []
                if normalized_term in source_text:
                    matched_columns.append("table")
                for column in searchable_columns:
                    value = getattr(r, column, "")
                    if value is not None and normalized_term in str(value).lower():
                        matched_columns.append(column)
                label = (
                    fields.get("client_name")
                    or fields.get("owner_name")
                    or fields.get("name")
                    or fields.get("full_name")
                    or fields.get("employee_id")
                    or fields.get("title")
                    or fields.get("property_code")
                    or fields.get("receipt_no")
                    or fields.get("invoice_no")
                    or fields.get("tenant_name")
                    or fields.get("vendor_name")
                    or ""
                )
                detail = (
                    fields.get("contact")
                    or fields.get("contact_phone")
                    or fields.get("owner_phone")
                    or fields.get("phone")
                    or fields.get("owner_contact")
                    or fields.get("email")
                    or fields.get("location")
                    or fields.get("area")
                    or fields.get("status")
                    or ""
                )
                results.append({
                    "table": search_table,
                    "source": source,
                    "id": r.id,
                    "label": str(label or ""),
                    "detail": str(detail or ""),
                    "matched_columns": matched_columns,
                    "fields": fields,
                })
        except Exception:
            pass
    return {"ok": True, "query": q, "count": len(results), "results": results}


@router.post("/ai-match")
def ai_match(
    req: MatchRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Explainable smart matching: find complementary deal records."""
    model = get_table_model(req.table)
    if model is None or req.table not in DEAL_TABLES:
        raise HTTPException(status_code=400, detail="AI matching is only supported for deal tables")
    require_table_read(user, req.table)
    record = db.query(model).filter(model.id == req.record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    if "requirements" in req.table:
        match_table_name = req.table.replace("requirements", "availability")
        match_model = get_table_model(match_table_name)
    elif "availability" in req.table:
        match_table_name = req.table.replace("availability", "requirements")
        match_model = get_table_model(match_table_name)
    else:
        return {"ok": True, "matches": []}
    if match_model is None:
        return {"ok": True, "matches": []}

    query = db.query(match_model)
    if hasattr(match_model, "status"):
        status_col = getattr(match_model, "status")
        query = query.filter(or_(status_col.is_(None), ~status_col.in_(["Rented", "Sold", "Closed", "Inactive"])))
    if hasattr(match_model, "workflow_stage"):
        stage_col = getattr(match_model, "workflow_stage")
        query = query.filter(or_(stage_col.is_(None), ~stage_col.in_(["Closed", "Deal Done"])))
    matches = best_matches(record, query.order_by(match_model.id.desc()).limit(500).all(), req.table, match_table_name, limit=20)
    return {"ok": True, "matches": matches}


@router.get("/pipeline/stats")
def pipeline_stats(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    counts = {stage: 0 for stage in DEAL_STAGES}
    totals = {stage: 0.0 for stage in DEAL_STAGES}
    for table in DEAL_TABLES:
        if not can_read_table(user, table):
            continue
        model = get_table_model(table)
        for record in db.query(model).all():
            row = deal_row(table, record)
            counts[row["stage"]] += 1
            totals[row["stage"]] += row["amount"]
    return {"ok": True, "stages": DEAL_STAGES, "counts": counts, "totals": totals}


@router.get("/pipeline/records")
def pipeline_records(
    stage: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if stage and stage not in DEAL_STAGES:
        raise HTTPException(status_code=400, detail=f"Invalid stage. Allowed: {DEAL_STAGES}")
    rows = []
    for table in DEAL_TABLES:
        if not can_read_table(user, table):
            continue
        model = get_table_model(table)
        query = db.query(model)
        if stage:
            query = query.filter(model.workflow_stage == stage)
        for record in query.order_by(model.id.desc()).limit(500).all():
            rows.append(deal_row(table, record))
    rows.sort(key=lambda r: (DEAL_STAGES.index(r["stage"]), r["next_follow_up"] or "9999-12-31"))
    return {"ok": True, "stage": stage or "All", "count": len(rows), "rows": rows}


@router.put("/{table}/{record_id}/workflow")
def update_workflow(
    table: str,
    record_id: int,
    req: WorkflowAction,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if table not in DEAL_TABLES:
        raise HTTPException(status_code=400, detail="Workflow is only supported for deal tables")
    require_table_write(user, table)
    if req.stage not in DEAL_STAGES:
        raise HTTPException(status_code=400, detail=f"Invalid stage. Allowed: {DEAL_STAGES}")
    if req.priority not in DEAL_PRIORITIES:
        raise HTTPException(status_code=400, detail=f"Invalid priority. Allowed: {DEAL_PRIORITIES}")
    model = get_table_model(table)
    record = db.query(model).filter(model.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    before = audit_snapshot(record)
    record.workflow_stage = req.stage
    record.priority = req.priority
    record.next_follow_up = req.next_follow_up
    record.assigned_to = req.assigned_to or user.username
    record.last_contacted = datetime.now().strftime("%Y-%m-%d")
    record.deal_probability = req.deal_probability if req.deal_probability is not None else STAGE_PROBABILITY[req.stage]
    if req.expected_close_value is not None:
        record.expected_close_value = req.expected_close_value
    record.lost_reason = req.lost_reason
    if req.status and hasattr(record, "status"):
        record.status = normalize_availability_status(req.status)
    elif req.stage == "Pending" and hasattr(record, "status"):
        record.status = "Pending"
    if req.stage in {"Closed", "Deal Done"}:
        record.closed_at = datetime.now()
        if req.status and hasattr(record, "status"):
            record.status = normalize_availability_status(req.status)
        elif table == "rent_availability" and hasattr(record, "status"):
            record.status = "Rented"
        elif table == "sale_availability" and hasattr(record, "status"):
            record.status = "Sold"
    sync_after_deal_save(db, table, record, user.username)
    write_audit_log(db, user, table, record_id, "workflow", before=before, after=audit_snapshot(record))
    db.commit()
    return {"ok": True, "message": "Workflow updated", "record": deal_row(table, record)}


@router.post("/duplicates/check")
def check_duplicates(
    payload: dict,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    table = str(payload.get("table") or "")
    if table not in ALLOWED_TABLES:
        raise HTTPException(status_code=400, detail="Invalid table")
    require_table_read(user, table)
    data = payload.get("data") or {}
    if not isinstance(data, dict):
        raise HTTPException(status_code=400, detail="data must be an object")
    record_id = payload.get("record_id")
    try:
        exclude_id = int(record_id) if record_id not in (None, "") else None
    except (TypeError, ValueError):
        exclude_id = None
    duplicates = find_duplicate_records(db, user, table, data, exclude_id=exclude_id)
    return {"ok": True, "duplicates": duplicates, "count": len(duplicates)}


@router.get("/followups/today")
def due_followups(
    limit: int = Query(120, le=500),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    today = date.today().isoformat()
    rows: list[dict] = []
    for table in DEAL_TABLES:
        if not can_read_table(user, table):
            continue
        model = get_table_model(table)
        if model is None:
            continue
        records = db.query(model).order_by(model.id.desc()).limit(1000).all()
        for record in records:
            stage = normalize_stage(getattr(record, "workflow_stage", None))
            if stage in {"Closed", "Deal Done"}:
                continue
            follow_up = str(getattr(record, "next_follow_up", "") or "").strip()
            priority = str(getattr(record, "priority", "") or "Medium")
            is_due = bool(follow_up and follow_up <= today)
            is_hot = priority in {"High", "Urgent"}
            if not is_due and not is_hot:
                continue
            row = deal_row(table, record)
            row["due_status"] = "Overdue" if follow_up and follow_up < today else "Due Today" if follow_up == today else "High Priority"
            row["label"] = TABLE_LABELS.get(table, table)
            rows.append(row)
    priority_order = {"Urgent": 0, "High": 1, "Medium": 2, "Low": 3}
    rows.sort(key=lambda r: (r.get("next_follow_up") or "9999-12-31", priority_order.get(r.get("priority"), 9)))
    return {"ok": True, "date": today, "count": len(rows[:limit]), "rows": rows[:limit]}


@router.get("/audit/logs")
def list_audit_logs(
    table: Optional[str] = Query(None),
    record_id: Optional[int] = Query(None),
    limit: int = Query(100, le=500),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    require_audit_permission(user)
    query = db.query(AuditLog)
    if table:
        query = query.filter(AuditLog.table_name == table)
    if record_id is not None:
        query = query.filter(AuditLog.record_id == record_id)
    total = query.count()
    logs = query.order_by(AuditLog.id.desc()).offset(offset).limit(limit).all()
    rows = []
    for item in logs:
        rows.append({
            "id": item.id,
            "table": item.table_name,
            "table_label": TABLE_LABELS.get(item.table_name, item.table_name),
            "record_id": item.record_id,
            "action": item.action,
            "username": item.username,
            "summary": item.summary,
            "created_at": item.created_at,
        })
    return {"ok": True, "total": total, "count": len(rows), "rows": rows}


@router.get("/backup/status")
def get_backup_status(user: User = Depends(require_permission("backup"))):
    return {"ok": True, **backup_status()}


@router.post("/backup/run")
def run_backup_now(user: User = Depends(require_permission("backup"))):
    path = run_database_backup("manual")
    return {"ok": True, "path": str(path), "message": "Backup created"}


PHASE1_IMPORT_ALIASES = {
    "rent_requirements": {
        "date": "date",
        "name": "client_name",
        "client name": "client_name",
        "contact": "contact_phone",
        "contact no": "contact_phone",
        "contact number": "contact_phone",
        "status": "client_status",
        "client broker owner": "client_status",
        "broker owner client": "client_status",
        "rooms": "size",
        "room": "size",
        "beds": "size",
        "bed": "size",
        "bedrooms": "size",
        "size": "measurement_unit",
        "size unit": "measurement_unit",
        "measurement unit": "measurement_unit",
        "area unit": "measurement_unit",
        "size type": "measurement_unit",
        "measurement": "measurement",
        "measurement value": "measurement",
        "area": "measurement",
        "unit": "measurement_unit",
        "sq ft yards": "measurement_unit",
        "sqr ft yards": "measurement_unit",
        "floor": "floor",
        "property requirement": "property_requires",
        "property type": "property_requires",
        "required property": "property_requires",
        "property required": "property_requires",
        "property needed": "property_requires",
        "property required needed": "property_requires",
        "needed property": "property_requires",
        "property requires": "property_requires",
        "location": "location",
        "family bachelor other": "bachelor_family",
        "family bachelor": "bachelor_family",
        "bachelor family": "bachelor_family",
        "persons": "persons",
        "facilities": "facilities",
        "budget": "budget",
    },
    "rent_availability": {
        "date": "date",
        "name": "owner_name",
        "owner name": "owner_name",
        "broker owner name": "owner_name",
        "contact": "owner_phone",
        "contact no": "owner_phone",
        "contact number": "owner_phone",
        "status": "client_broker",
        "client broker owner": "client_broker",
        "broker owner client": "client_broker",
        "rooms": "size",
        "room": "size",
        "beds": "size",
        "bed": "size",
        "bedrooms": "size",
        "size": "measurement_unit",
        "size unit": "measurement_unit",
        "measurement unit": "measurement_unit",
        "area unit": "measurement_unit",
        "size type": "measurement_unit",
        "measurement": "measurement",
        "measurement value": "measurement",
        "area": "measurement",
        "unit": "measurement_unit",
        "sq ft yards": "measurement_unit",
        "sqr ft yards": "measurement_unit",
        "floor": "floor",
        "property availability": "property_availability",
        "property type": "property_availability",
        "property available": "property_availability",
        "available property": "property_availability",
        "rent": "monthly_rent",
        "monthly rent": "monthly_rent",
        "advance": "deposit",
        "deposit": "deposit",
        "maintenance": "maintenance_charge",
        "maintainance": "maintenance_charge",
        "location": "location",
        "building name": "building_name",
        "family bachelor other": "bachelor_family",
        "family bachelor": "bachelor_family",
        "bachelor family": "bachelor_family",
        "persons": "persons",
        "facilities": "facilities",
    },
    "sale_requirements": {
        "date": "date",
        "name": "client_name",
        "client name": "client_name",
        "contact": "contact_phone",
        "contact no": "contact_phone",
        "contact number": "contact_phone",
        "status": "client_status",
        "client broker owner": "client_status",
        "broker owner client": "client_status",
        "rooms": "size",
        "room": "size",
        "beds": "size",
        "bed": "size",
        "bedrooms": "size",
        "size": "measurement_unit",
        "size unit": "measurement_unit",
        "measurement unit": "measurement_unit",
        "area unit": "measurement_unit",
        "size type": "measurement_unit",
        "measurement": "measurement",
        "measurement value": "measurement",
        "area": "measurement",
        "unit": "measurement_unit",
        "sq ft yards": "measurement_unit",
        "sqr ft yards": "measurement_unit",
        "floor": "floor",
        "budget": "budget",
        "maintenance": "maintenance_charge",
        "maintainance": "maintenance_charge",
        "location": "location",
        "required property": "property_requires",
        "property requirement": "property_requires",
        "property type": "property_requires",
        "property required": "property_requires",
        "property needed": "property_requires",
        "property required needed": "property_requires",
        "needed property": "property_requires",
        "family bachelor other": "bachelor_family",
        "family bachelor": "bachelor_family",
        "bachelor family": "bachelor_family",
        "facilities": "facilities",
    },
    "sale_availability": {
        "date": "date",
        "name": "owner_name",
        "owner name": "owner_name",
        "broker owner name": "owner_name",
        "contact": "owner_phone",
        "contact no": "owner_phone",
        "contact number": "owner_phone",
        "status": "client_broker",
        "client broker owner": "client_broker",
        "broker owner client": "client_broker",
        "rooms": "size",
        "room": "size",
        "beds": "size",
        "bed": "size",
        "bedrooms": "size",
        "size": "measurement_unit",
        "size unit": "measurement_unit",
        "measurement unit": "measurement_unit",
        "area unit": "measurement_unit",
        "size type": "measurement_unit",
        "measurement": "measurement",
        "measurement value": "measurement",
        "area": "measurement",
        "unit": "measurement_unit",
        "sq ft yards": "measurement_unit",
        "sqr ft yards": "measurement_unit",
        "floor": "floor",
        "demand": "demand",
        "maintenance": "maintenance_charge",
        "maintainance": "maintenance_charge",
        "location": "location",
        "building name": "building_name",
        "property availability": "property_availability",
        "property type": "property_availability",
        "property available": "property_availability",
        "available property": "property_availability",
        "facilities": "facilities",
    },
}


def normalize_import_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()


def import_cell_text(value: object) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return ""
    return str(value).strip()


def import_records_from_grid(table: str, grid: list[list[object]]) -> list[dict]:
    rows = [row for row in grid if any(str(cell or "").strip() for cell in row)]
    if len(rows) < 2:
        return []
    model = get_table_model(table)
    columns = model_columns(model) if model is not None else set()
    aliases = PHASE1_IMPORT_ALIASES[table]
    headers = [normalize_import_header(cell) for cell in rows[0]]
    keys = []
    for header in headers:
        key = aliases.get(header) or aliases.get(header.replace(" no", ""))
        if key is None and header in columns:
            key = header
        keys.append(key)
    records: list[dict] = []
    for row in rows[1:]:
        record: dict = {}
        for index, value in enumerate(row):
            key = keys[index] if index < len(keys) else None
            if not key:
                continue
            text_value = import_cell_text(value)
            if text_value != "":
                record[key] = text_value
        if any(str(value or "").strip() for value in record.values()):
            records.append(record)
    return records


def decode_import_content(payload: dict) -> bytes:
    encoded = str(payload.get("content_base64") or "")
    if "," in encoded and encoded.lower().startswith("data:"):
        encoded = encoded.split(",", 1)[1]
    if not encoded:
        raise HTTPException(status_code=400, detail="No import file content received")
    try:
        return base64.b64decode(encoded)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Could not read import file") from exc


@router.post("/{table}/import/preview")
def preview_table_import(
    table: str,
    payload: dict,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if table not in PHASE1_TABLES:
        raise HTTPException(status_code=400, detail="Import is available for Phase 1 tables")
    require_table_write(user, table)
    filename = str(payload.get("filename") or "").lower()
    raw = decode_import_content(payload)
    if filename.endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise HTTPException(status_code=500, detail="Excel import needs openpyxl installed") from exc
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        grid = [list(row) for row in sheet.iter_rows(values_only=True)]
    elif filename.endswith(".xls"):
        raise HTTPException(status_code=400, detail="Please save old .xls files as .xlsx or CSV before import")
    else:
        try:
            text_value = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text_value = raw.decode("cp1252")
        grid = [row for row in csv.reader(io.StringIO(text_value))]
    records = import_records_from_grid(table, grid)
    return {"ok": True, "table": table, "count": len(records), "records": records}


@router.get("/phase1/settings")
def get_phase1_settings(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    return {
        "ok": True,
        "company_name": setting_value(db, "company_name", "MBM Enterprises"),
        "company_address": setting_value(db, "company_address", ""),
        "company_phone": setting_value(db, "company_phone", ""),
        "company_email": setting_value(db, "company_email", ""),
        "company_logo": setting_value(db, "company_logo", ""),
        "currency_symbol": setting_value(db, "currency_symbol", "Rs."),
        "default_commission": setting_value(db, "default_commission", ""),
        "tax_rate": setting_value(db, "tax_rate", ""),
        "bank_account": setting_value(db, "bank_account", ""),
        "theme": setting_value(db, "phase1_theme", "Light"),
        "areas": setting_list(db, "phase1_areas", PHASE1_AREAS),
        "facilities": setting_list(db, "phase1_facilities", PHASE1_FACILITIES),
        "floors": setting_list(db, "phase1_floors", PHASE1_FLOORS),
        "property_types": setting_list(db, "phase1_property_types", PHASE1_PROPERTY_TYPES),
        "measurement_units": setting_list(db, "phase1_measurement_units", PHASE1_MEASUREMENT_UNITS),
        "expense_categories": setting_list(db, "expense_categories", list(EXPENSE_CATEGORIES)),
    }


@router.put("/phase1/settings")
def save_phase1_settings(payload: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_admin(user):
        raise HTTPException(status_code=403, detail="Only admins can change settings")
    for key in (
        "company_name", "company_address", "company_phone", "company_email",
        "company_logo", "currency_symbol", "default_commission", "tax_rate", "bank_account",
    ):
        set_setting(db, key, str(payload.get(key) or ""))
    set_setting(db, "phase1_theme", str(payload.get("theme") or "Light"))
    list_settings = {
        "areas": ("phase1_areas", PHASE1_AREAS),
        "facilities": ("phase1_facilities", PHASE1_FACILITIES),
        "floors": ("phase1_floors", PHASE1_FLOORS),
        "property_types": ("phase1_property_types", PHASE1_PROPERTY_TYPES),
        "measurement_units": ("phase1_measurement_units", PHASE1_MEASUREMENT_UNITS),
        "expense_categories": ("expense_categories", list(EXPENSE_CATEGORIES)),
    }
    for payload_key, (setting_key, defaults) in list_settings.items():
        if payload_key not in payload:
            continue
        set_setting(db, setting_key, setting_list_text(payload.get(payload_key), defaults))
    db.commit()
    return {"ok": True, "message": "Settings saved"}


@router.get("/options")
def record_options(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    expense_categories = setting_list(db, "expense_categories", list(EXPENSE_CATEGORIES))
    return {
        "ok": True,
        "expense_categories": expense_categories,
        "tables": {
            "expense_transactions": {
                "expense_category": expense_categories,
            },
        },
    }


@router.get("/approvals/pending")
def list_pending_approvals(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if not is_admin(user):
        raise HTTPException(status_code=403, detail="Only admins can view approvals")
    rows = db.query(PendingApproval).filter(PendingApproval.status == "Pending").order_by(PendingApproval.id.desc()).all()
    return {
        "ok": True,
        "rows": [
            {
                "id": row.id,
                "action": row.action,
                "table_name": row.table_name,
                "table_label": TABLE_LABELS.get(row.table_name, row.table_name),
                "record_id": row.record_id,
                "payload": json.loads(row.payload or "{}"),
                "requested_by": row.requested_by,
                "requested_at": row.requested_at,
                "status": row.status,
            }
            for row in rows
        ],
    }


@router.post("/approvals/{approval_id}/review")
def review_pending_approval(
    approval_id: int,
    payload: dict,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if not is_admin(user):
        raise HTTPException(status_code=403, detail="Only admins can review approvals")
    approval = db.query(PendingApproval).filter(PendingApproval.id == approval_id).first()
    if not approval or approval.status != "Pending":
        raise HTTPException(status_code=404, detail="Pending approval not found")
    approved = bool(payload.get("approved"))
    comment = str(payload.get("comment") or "")
    if approved:
        model = get_table_model(approval.table_name)
        if model is None:
            raise HTTPException(status_code=400, detail="Unsupported table")
        record = db.query(model).filter(model.id == approval.record_id).first()
        if not record:
            raise HTTPException(status_code=404, detail="Record not found")
        before = audit_snapshot(record)
        data = json.loads(approval.payload or "{}")
        if approval.action == "edit":
            data = normalize_phase1_payload(approval.table_name, data, model_columns(model))
            validate_record_payload(approval.table_name, data, creating=False)
            data["last_edited_by"] = user.username
            data["last_edited_at"] = datetime.now().isoformat(timespec="seconds")
            for key, value in data.items():
                if hasattr(record, key) and key != "id":
                    setattr(record, key, value)
            sync_after_deal_save(db, approval.table_name, record, user.username)
            write_audit_log(db, user, approval.table_name, approval.record_id, "update", before=before, after=audit_snapshot(record))
        elif approval.action == "delete":
            if hasattr(record, "is_deleted"):
                record.is_deleted = True
                record.deleted_by = approval.requested_by
                record.deleted_at = datetime.now().isoformat(timespec="seconds")
                write_audit_log(db, user, approval.table_name, approval.record_id, "delete", before=before, after=audit_snapshot(record))
        elif approval.action == "restore":
            if hasattr(record, "is_deleted"):
                record.is_deleted = False
                record.deleted_by = None
                record.deleted_at = None
                write_audit_log(db, user, approval.table_name, approval.record_id, "restore", before=before, after=audit_snapshot(record))
    approval.status = "Approved" if approved else "Rejected"
    approval.reviewed_by = user.username
    approval.reviewed_at = datetime.now().isoformat(timespec="seconds")
    approval.review_comment = comment
    db.commit()
    return {"ok": True, "message": approval.status}


@router.post("/match")
def phase1_match(payload: dict, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    table = str(payload.get("table") or "")
    record_id = int(payload.get("record_id") or 0)
    if table not in {"rent_requirements", "sale_requirements"}:
        raise HTTPException(status_code=400, detail="Match starts from a rent or sale requirement")
    require_table_read(user, table)
    target_table = "rent_availability" if table == "rent_requirements" else "sale_availability"
    require_table_read(user, target_table)
    source_model = get_table_model(table)
    target_model = get_table_model(target_table)
    requirement = db.query(source_model).filter(source_model.id == record_id).first()
    if not requirement:
        raise HTTPException(status_code=404, detail="Requirement not found")
    facilities = setting_list(db, "phase1_facilities", PHASE1_FACILITIES)
    rows = db.query(target_model).filter(getattr(target_model, "is_deleted").is_(False) | (getattr(target_model, "is_deleted") == 0) | getattr(target_model, "is_deleted").is_(None)).order_by(target_model.id.desc()).limit(2000).all()
    results = []
    req_location = normalize_text(getattr(requirement, "location", ""))
    req_amount = float(getattr(requirement, "budget", 0) or 0)
    req_rooms = normalize_text(getattr(requirement, "size", ""))
    req_floors = multi_value_tokens(getattr(requirement, "floor", ""))
    req_facilities = parse_facilities(getattr(requirement, "facilities", ""), facilities)
    for row in rows:
        score = 0.0
        reasons: list[str] = []
        row_location = normalize_text(getattr(row, "location", ""))
        if req_location and row_location:
            if req_location == row_location:
                score += 45
                reasons.append("same location")
            elif nearby_location(req_location, row_location):
                score += 35
                reasons.append("nearby/similar location")
            elif req_location in row_location or row_location in req_location:
                score += 25
                reasons.append("similar location text")
        amount_key = "monthly_rent" if target_table == "rent_availability" else "demand"
        row_amount = float(getattr(row, amount_key, 0) or 0)
        if req_amount and row_amount:
            if row_amount <= req_amount:
                score += 20
                reasons.append("price within budget")
            elif row_amount <= req_amount * 1.1:
                score += 12
                reasons.append("price near budget")
        if req_rooms and req_rooms == normalize_text(getattr(row, "size", "")):
            score += 15
            reasons.append("matching rooms")
        if req_floors and req_floors & multi_value_tokens(getattr(row, "floor", "")):
            score += 10
            reasons.append("matching floor")
        overlap = req_facilities & parse_facilities(getattr(row, "facilities", ""), facilities)
        if overlap:
            score += min(10, len(overlap) * 3)
            reasons.append(f"{len(overlap)} facilities matched")
        if score > 0:
            record = serialize_record(row)
            results.append({
                "id": row.id,
                "score": min(score, 100),
                "reasons": reasons,
                "record": record,
                "name": getattr(row, "owner_name", ""),
                "contact": record_text(row, "owner_phone", "contact_phone", "contact"),
                "location": getattr(row, "location", ""),
                "rooms": getattr(row, "size", ""),
                "floor": getattr(row, "floor", ""),
                "amount": row_amount,
            })
    results.sort(key=lambda item: item["score"], reverse=True)
    return {"ok": True, "target_table": target_table, "requirement": serialize_record(requirement), "matches": results[:50]}


@router.get("/ecosystem/health")
def ecosystem_health(user: User = Depends(get_current_user)):
    if normalize_role(user.role) not in {"Super Admin", "Admin", "Manager"}:
        raise HTTPException(status_code=403, detail="Only managers and admins can view ecosystem health")
    return collect_ecosystem_health()


@router.get("/{table}/template.csv")
def table_template(table: str, user: User = Depends(get_current_user)):
    if table not in PHASE1_TABLES:
        raise HTTPException(status_code=400, detail="Template is available for Phase 1 tables")
    headers = {
        "rent_requirements": ["Name", "Status", "Contact", "Date", "Property Required / Needed", "Rooms", "Measurement", "Size", "Floor", "Location", "Family / Bachelor / Other", "Persons", "Facilities", "Budget"],
        "rent_availability": ["Name", "Status", "Contact", "Date", "Property Available", "Rooms", "Measurement", "Size", "Floor", "Rent", "Advance", "Maintenance", "Location", "Building Name", "Family / Bachelor / Other", "Persons", "Facilities"],
        "sale_requirements": ["Name", "Status", "Contact", "Date", "Property Required / Needed", "Rooms", "Measurement", "Size", "Floor", "Budget", "Maintenance", "Location", "Family / Bachelor / Other", "Facilities"],
        "sale_availability": ["Name", "Status", "Contact", "Date", "Property Available", "Rooms", "Measurement", "Size", "Floor", "Demand", "Maintenance", "Location", "Building Name", "Facilities"],
    }[table]
    return Response(",".join(headers) + "\n", media_type="text/csv", headers={"Content-Disposition": f"attachment; filename={table}_template.csv"})


@router.get("/{table}")
def list_records(
    table: str,
    limit: int = Query(100, le=PHASE1_LIST_LIMIT),
    offset: int = Query(0, ge=0),
    q: str = Query(""),
    keyword: Optional[str] = Query(None),
    stage: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    area: Optional[str] = Query(None),
    office_address: Optional[str] = Query(None),
    home_address: Optional[str] = Query(None),
    sort: Optional[str] = Query(None),
    sort_by: Optional[str] = Query(None),
    direction: str = Query("desc"),
    sort_order: Optional[str] = Query(None),
    include_deleted: bool = Query(False),
    deleted_only: bool = Query(False),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if table not in ALLOWED_TABLES:
        raise HTTPException(status_code=400, detail=f"Invalid table. Allowed: {sorted(ALLOWED_TABLES)}")
    require_table_read(user, table)
    model = get_table_model(table)
    if model is None:
        raise HTTPException(status_code=400, detail="Table not supported")
    if (include_deleted or deleted_only) and (table not in PHASE1_TABLES or not is_admin(user)):
        raise HTTPException(status_code=403, detail="Only admins can view recycled records")
    keyword_value = keyword if keyword not in (None, "") else q
    date_from_value = date_from or start_date
    date_to_value = date_to or end_date
    field_filters = {}
    if table == "broker_contacts":
        field_filters = {
            "area": area,
            "office_address": office_address,
            "home_address": home_address,
        }
    query = apply_list_filters(
        db.query(model),
        model,
        table,
        q=keyword_value,
        stage=stage,
        status=status,
        date_from=date_from_value,
        date_to=date_to_value,
        field_filters=field_filters,
        include_deleted=include_deleted,
        deleted_only=deleted_only,
    )
    total = query.count()
    effective_limit = limit
    effective_offset = offset
    if table in PHASE1_TABLES and not deleted_only:
        # Phase 1 deal desk lists are office working lists, not paged archives.
        # Return the complete active set even if an already-open browser tab is
        # still running older JavaScript that asks for only 100 rows or page 2.
        effective_limit = max(limit, min(total, PHASE1_LIST_LIMIT))
        effective_offset = 0
    default_sort, default_direction = DEFAULT_TABLE_SORTS.get(table, ("id", "desc"))
    requested_sort = (sort_by or sort or default_sort).strip()
    valid_columns = model_columns(model)
    if requested_sort not in valid_columns:
        raise HTTPException(status_code=400, detail=f"Invalid sort_by: {requested_sort}")
    direction_source = sort_order or direction
    if not (sort_by or sort or sort_order) and direction == "desc":
        direction_source = default_direction
    direction_value = (direction_source or default_direction).strip().lower()
    if direction_value not in SORT_DIRECTIONS:
        raise HTTPException(status_code=400, detail="Invalid sort direction. Use asc or desc.")
    sort_column = getattr(model, requested_sort)
    order = sort_column.asc() if direction_value == "asc" else sort_column.desc()
    orders = [order]
    if requested_sort != "id" and "id" in valid_columns:
        orders.append(getattr(model, "id").desc())
    rows = query.order_by(*orders).offset(effective_offset).limit(effective_limit).all()
    return {"ok": True, "table": table, "total": total, "count": len(rows), "rows": [serialize_record(r) for r in rows]}


@router.get("/{table}/{record_id}")
def get_record(
    table: str,
    record_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if table not in ALLOWED_TABLES:
        raise HTTPException(status_code=400, detail="Invalid table")
    require_table_read(user, table)
    model = get_table_model(table)
    record = db.query(model).filter(model.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    return {"ok": True, "record": serialize_record(record)}


@router.post("/{table}")
def create_record(
    table: str,
    req: RecordCreate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if table not in ALLOWED_TABLES:
        raise HTTPException(status_code=400, detail="Invalid table")
    require_table_write(user, table)
    model = get_table_model(table)
    if model is None:
        raise HTTPException(status_code=400, detail="Table not supported")
    data = clean_payload(model, req.data, user=user, creating=True, table=table)
    if not data:
        raise HTTPException(status_code=400, detail="No valid fields to create")
    validate_record_payload(table, data, creating=True)
    duplicates = find_duplicate_records(db, user, table, data)
    instance = model(**data)
    db.add(instance)
    db.flush()
    sync_after_deal_save(db, table, instance, user.username)
    write_audit_log(db, user, table, instance.id, "create", after=audit_snapshot(instance))
    db.commit()
    db.refresh(instance)
    return {"ok": True, "id": instance.id, "message": "Record created", "duplicates": duplicates}


@router.put("/{table}/{record_id}")
def update_record(
    table: str,
    record_id: int,
    req: RecordUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if table not in ALLOWED_TABLES:
        raise HTTPException(status_code=400, detail="Invalid table")
    require_table_write(user, table)
    model = get_table_model(table)
    record = db.query(model).filter(model.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    data = clean_payload(model, req.data, table=table)
    if not data:
        raise HTTPException(status_code=400, detail="No valid fields to update")
    if table == "attendance":
        columns = set(model_column_map(model))
        merged = {key: getattr(record, key, None) for key in columns}
        merged.update(data)
        data = {key: value for key, value in calculate_attendance(merged).items() if key in columns}
    validate_record_payload(table, data, creating=False)
    if table in PHASE1_TABLES and not is_admin(user):
        db.add(PendingApproval(
            action="edit",
            table_name=table,
            record_id=record_id,
            payload=json.dumps(data, default=str),
            requested_by=user.username,
            requested_at=datetime.now().isoformat(timespec="seconds"),
            status="Pending",
        ))
        db.commit()
        return {"ok": True, "message": "Edit sent for admin approval", "pending": True}
    before = audit_snapshot(record)
    if table in PHASE1_TABLES:
        data["last_edited_by"] = user.username
        data["last_edited_at"] = datetime.now().isoformat(timespec="seconds")
    for key, value in data.items():
        setattr(record, key, value)
    sync_after_deal_save(db, table, record, user.username)
    write_audit_log(db, user, table, record_id, "update", before=before, after=audit_snapshot(record))
    db.commit()
    return {"ok": True, "message": "Record updated"}


@router.delete("/{table}/{record_id}")
def delete_record(
    table: str,
    record_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if table not in ALLOWED_TABLES:
        raise HTTPException(status_code=400, detail="Invalid table")
    require_table_write(user, table)
    model = get_table_model(table)
    record = db.query(model).filter(model.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    if table in PHASE1_TABLES and not is_admin(user):
        require_table_write(user, table)
        db.add(PendingApproval(
            action="delete",
            table_name=table,
            record_id=record_id,
            payload="{}",
            requested_by=user.username,
            requested_at=datetime.now().isoformat(timespec="seconds"),
            status="Pending",
        ))
        db.commit()
        return {"ok": True, "message": "Delete sent for admin approval", "pending": True}
    if table not in PHASE1_TABLES and not has_permission(user.role, "delete"):
        raise HTTPException(status_code=403, detail="Permission 'delete' required")
    require_no_child_references(db, table, record_id)
    before = audit_snapshot(record)
    if table in PHASE1_TABLES and hasattr(record, "is_deleted"):
        record.is_deleted = True
        record.deleted_by = user.username
        record.deleted_at = datetime.now().isoformat(timespec="seconds")
        write_audit_log(db, user, table, record_id, "delete", before=before, after=audit_snapshot(record))
    else:
        write_audit_log(db, user, table, record_id, "delete", before=before)
        db.delete(record)
    db.commit()
    return {"ok": True, "message": "Record recycled" if table in PHASE1_TABLES else "Record deleted"}


@router.post("/{table}/{record_id}/restore")
def restore_record(
    table: str,
    record_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if table not in PHASE1_TABLES:
        raise HTTPException(status_code=400, detail="Restore is only available for Phase 1 tables")
    if not is_admin(user):
        raise HTTPException(status_code=403, detail="Only admins can restore records")
    model = get_table_model(table)
    record = db.query(model).filter(model.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    before = audit_snapshot(record)
    record.is_deleted = False
    record.deleted_by = None
    record.deleted_at = None
    write_audit_log(db, user, table, record_id, "restore", before=before, after=audit_snapshot(record))
    db.commit()
    return {"ok": True, "message": "Record restored"}


@router.put("/{table}/{record_id}/approve")
def approve_record(
    table: str,
    record_id: int,
    req: ApprovalAction,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if user.role not in ("Super Admin", "Admin"):
        raise HTTPException(status_code=403, detail="Only admins can approve")
    if table not in ALLOWED_TABLES:
        raise HTTPException(status_code=400, detail="Invalid table")
    require_table_write(user, table)
    model = get_table_model(table)
    record = db.query(model).filter(model.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    if not hasattr(record, "approval_status"):
        raise HTTPException(status_code=400, detail="This table does not support approvals")
    before = audit_snapshot(record)
    record.approval_status = req.status
    record.approval_comment = req.comment
    record.approved_by = user.username
    record.approved_at = datetime.now()
    write_audit_log(db, user, table, record_id, "approve", before=before, after=audit_snapshot(record))
    db.commit()
    return {"ok": True, "message": f"Record #{record_id} marked as {req.status}"}


# (ai_match and global_search are defined above, before parameterized routes)
